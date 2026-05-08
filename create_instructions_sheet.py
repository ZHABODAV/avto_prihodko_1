"""
Create Instructions sheet in Excel template with comprehensive user guide in Russian.
This script adds a formatted Instructions sheet to the terminal_template.xlsx file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_instructions_sheet():
    """Add Instructions sheet to terminal_template.xlsx"""
    
    # Load existing workbook
    wb = openpyxl.load_workbook('terminal_template.xlsx')
    
    # Check if Instructions sheet already exists, if so remove it
    if 'Instructions' in wb.sheetnames:
        del wb['Instructions']
    
    # Create new Instructions sheet at the beginning
    ws = wb.create_sheet('Instructions', 0)
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    header_font = Font(name='Arial', size=14, bold=True, color='1F4E78')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    subheader_font = Font(name='Arial', size=12, bold=True, color='305496')
    
    body_font = Font(name='Arial', size=11)
    example_font = Font(name='Arial', size=10, italic=True, color='7F7F7F')
    example_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    warning_font = Font(name='Arial', size=11, bold=True, color='C00000')
    warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    tip_font = Font(name='Arial', size=11, color='375623')
    tip_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 5
    
    row = 1
    
    # Title
    ws.merge_cells(f'B{row}:B{row}')
    cell = ws[f'B{row}']
    cell.value = 'ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ TERMINAL OPTIMIZER'
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 30
    row += 2
    
    # Section 1: Быстрый старт
    ws[f'B{row}'] = '1. БЫСТРЫЙ СТАРТ'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    steps = [
        '1. Откройте файл terminal_template.xlsx',
        '2. Перейдите на лист Input_Vessels',
        '3. Заполните данные минимум об одном судне (Vessel ID, ETA, Cargo Type, Volume, Priority)',
        '4. Заполните остальные входные листы (Input_Rail, Input_Demand, Input_CurrentState)',
        '5. Нажмите кнопку "Calculate Plan" на панели инструментов',
        '6. Дождитесь завершения расчета (появится окно с процессом)',
        '7. Просмотрите результаты на листах Output_* (Schedule, Gantt, Railway, Costs, Balance, Risks)',
        '8. При необходимости экспортируйте результаты в PDF используя кнопку "Export PDF"'
    ]
    
    for step in steps:
        ws[f'B{row}'] = step
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = left_align
        row += 1
    
    row += 1
    
    # Section 2: Входные данные
    ws[f'B{row}'] = '2. ВХОДНЫЕ ДАННЫЕ'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Input_Vessels
    ws[f'B{row}'] = '2.1 Input_Vessels (Информация о судах)'
    ws[f'B{row}'].font = subheader_font
    row += 1
    
    vessel_fields = [
        'Vessel ID: уникальный идентификатор судна (например: VICTORIA, OLEANDER)',
        'ETA: ожидаемое время прибытия (формат: ГГГГ-ММ-ДД ЧЧ:ММ или ДД.ММ.ГГГГ)',
        'Cargo Type: тип груза (sunflower = подсолнечник, palm = пальма)',
        'Total Volume: общий объем груза в тоннах',
        'Priority: приоритет (HIGH/MEDIUM/LOW)',
        'Demurrage Rate: ставка демереджа в USD/час (необязательно, по умолчанию 800)',
        'Strategic Weight: стратегический вес клиента 0-1 (необязательно, по умолчанию 1.0)'
    ]
    
    for field in vessel_fields:
        ws[f'B{row}'] = '  • ' + field
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = left_align
        row += 1
    
    ws[f'B{row}'] = 'Пример: VICTORIA | 2025-01-15 08:00 | sunflower | 35000 | HIGH | 800 | 1.0'
    ws[f'B{row}'].font = example_font
    ws[f'B{row}'].fill = example_fill
    row += 2
    
    # Input_Rail
    ws[f'B{row}'] = '2.2 Input_Rail (График подач вагонов с ЖДЦ)'
    ws[f'B{row}'].font = subheader_font
    row += 1
    
    ws[f'B{row}'] = '  • Date: дата подачи\n  • Time: время подачи\n  • Product: тип продукта\n  • Volume: объем в тоннах\n  • Wagons: количество вагонов'
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 75
    row += 2
    
    # Input_Demand
    ws[f'B{row}'] = '2.3 Input_Demand (Заявки клиентов)'
    ws[f'B{row}'].font = subheader_font
    row += 1
    
    ws[f'B{row}'] = '  • Week: номер недели\n  • Client: название клиента\n  • Product: продукт\n  • Volume: объем заявки в тоннах\n  • Priority: приоритет клиента'
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 75
    row += 2
    
    # Input_CurrentState
    ws[f'B{row}'] = '2.4 Input_CurrentState (Текущее состояние РВС)'
    ws[f'B{row}'].font = subheader_font
    row += 1
    
    ws[f'B{row}'] = '  • Tank ID: номер резервуара (РВС-1, РВС-2, РВС-3, РВС-4, РВС-5, ШТ)\n  • Product: текущий продукт в резервуаре\n  • Volume: текущий объем в тоннах\n  • State: состояние (idle/filling/discharging/analysis/cleaning)'
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 60
    row += 2
    
    # Config_Parameters
    ws[f'B{row}'] = '2.5 Config_Parameters (Параметры расчета)'
    ws[f'B{row}'].font = subheader_font
    row += 1
    
    ws[f'B{row}'] = '  • Horizon Days: горизонт планирования в днях (обычно 7-30)\n  • Rail Rate: скорость разгрузки ж/д, т/час\n  • Ship Load Rate: скорость погрузки судна, т/час\n  • Safety Margin: запас безопасности (обычно 0.10 = 10%)'
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 60
    row += 2
    
    # Section 3: Результаты
    ws[f'B{row}'] = '3. РЕЗУЛЬТАТЫ РАСЧЕТА'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    outputs = [
        ('Output_Schedule', 'График всех операций: швартовка, слив/налив, перекачка, зачистка'),
        ('Output_Gantt', 'Визуальная диаграмма Ганта с ресурсами (РВС, причал, ж/д платформа)'),
        ('Output_Railway', 'План ж/д диспетчера: подачи вагонов по сменам и дням'),
        ('Output_Costs', 'Расчет затрат: демередж, зачистка, аренда вагонов, итого'),
        ('Output_Balance', 'Балансы РВС по часам: приход, расход, остаток'),
        ('Output_Risks', 'Прогноз рисков и рекомендации по их минимизации')
    ]
    
    for name, desc in outputs:
        ws[f'B{row}'] = f'  • {name}: {desc}'
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = left_align
        ws.row_dimensions[row].height = 30
        row += 1
    
    row += 1
    
    # Section 4: Использование VBA макросов
    ws[f'B{row}'] = '4. ИСПОЛЬЗОВАНИЕ МАКРОСОВ'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    macros = [
        ('Calculate Plan', 'Запуск расчета оптимизации (требуется Python 3.9+)'),
        ('Clear Results', 'Очистка всех листов с результатами'),
        ('Export PDF', 'Экспорт результатов в PDF файл'),
        ('Help', 'Показать эту инструкцию')
    ]
    
    for macro, desc in macros:
        ws[f'B{row}'] = f'  • {macro}: {desc}'
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = left_align
        row += 1
    
    row += 2
    
    # Section 5: Устранение проблем
    ws[f'B{row}'] = '5. УСТРАНЕНИЕ ПРОБЛЕМ'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    ws[f'B{row}'] = '⚠️ ПРОБЛЕМА: Python не найден'
    ws[f'B{row}'].font = warning_font
    ws[f'B{row}'].fill = warning_fill
    ws[f'B{row}'].alignment = left_align
    row += 1
    
    ws[f'B{row}'] = '✓ РЕШЕНИЕ: Установите Python 3.9 или выше с https://www.python.org/downloads/\nПри установке обязательно отметьте "Add Python to PATH".\nЗатем запустите install.bat для установки зависимостей.'
    ws[f'B{row}'].font = tip_font
    ws[f'B{row}'].fill = tip_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 50
    row += 2
    
    ws[f'B{row}'] = '⚠️ ПРОБЛЕМА: Ошибка валидации входных данных'
    ws[f'B{row}'].font = warning_font
    ws[f'B{row}'].fill = warning_fill
    ws[f'B{row}'].alignment = left_align
    row += 1
    
    ws[f'B{row}'] = '✓ РЕШЕНИЕ: Проверьте:\n  • Все обязательные поля заполнены\n  • Даты в правильном формате\n  • Объемы положительные числа\n  • Типы продуктов правильные (sunflower/palm)'
    ws[f'B{row}'].font = tip_font
    ws[f'B{row}'].fill = tip_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 65
    row += 2
    
    ws[f'B{row}'] = '⚠️ ПРОБЛЕМА: Ошибка при выполнении расчета'
    ws[f'B{row}'].font = warning_font
    ws[f'B{row}'].fill = warning_fill
    ws[f'B{row}'].alignment = left_align
    row += 1
    
    ws[f'B{row}'] = '✓ РЕШЕНИЕ: Проверьте лог-файлы в папке logs/ для детальной информации об ошибке.\nЧастые причины:\n  • Недостаточно емкости РВС\n  • Несовместимые продукты без зачистки\n  • Превышение лимита вагонов'
    ws[f'B{row}'].font = tip_font
    ws[f'B{row}'].fill = tip_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 75
    row += 2
    
    ws[f'B{row}'] = '⚠️ ПРОБЛЕМА: Долгое выполнение расчета'
    ws[f'B{row}'].font = warning_font
    ws[f'B{row}'].fill = warning_fill
    ws[f'B{row}'].alignment = left_align
    row += 1
    
    ws[f'B{row}'] = '✓ РЕШЕНИЕ: Это нормально для сложных сценариев (>5 судов, >14 дней).\nОжидаемое время:\n  • Простой сценарий (1-2 судна, 7 дней): <10 сек\n  • Средний сценарий (3-5 судов, 14 дней): <30 сек\n  • Сложный сценарий (>5 судов, 30 дней): <2 минуты'
    ws[f'B{row}'].font = tip_font
    ws[f'B{row}'].fill = tip_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 90
    row += 2
    
    # Section 6: Контакты поддержки
    ws[f'B{row}'] = '6. КОНТАКТЫ ПОДДЕРЖКИ'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    ws[f'B{row}'] = 'При возникновении проблем обратитесь в службу поддержки:\n\nEmail: support@terminal-optimizer.local\nВнутренний телефон: доб. 1234\n\nДокументация: см. папку docs/ для подробных технических руководств'
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].alignment = left_align
    ws.row_dimensions[row].height = 75
    row += 2
    
    # Footer
    ws[f'B{row}'] = 'Terminal Optimizer v1.0.0 | © 2025 | Последнее обновление: Декабрь 2025'
    ws[f'B{row}'].font = Font(name='Arial', size=9, italic=True, color='7F7F7F')
    ws[f'B{row}'].alignment = center_align
    
    # Freeze first row
    ws.freeze_panes = 'B2'
    
    # Save workbook
    wb.save('terminal_template.xlsx')
    print("OK: Instructions sheet created successfully in terminal_template.xlsx")

if __name__ == "__main__":
    create_instructions_sheet()
