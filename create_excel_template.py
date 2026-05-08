"""
Terminal Optimizer - Excel Template Creator
Creates a properly formatted Excel template with sample data
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os

def create_template_with_sample_data(filename="terminal_template.xlsx"):
    """Create Excel template with sample data for quick start"""
    
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    example_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================
    # Sheet 1: Input_Vessels
    # ========================================
    ws_vessels = wb.create_sheet("Input_Vessels")
    
    headers = ["VesselID", "ETA", "CargoType", "TotalVolume", "DemurrageRate", "Priority", "Active", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_vessels.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Sample data
    sample_vessels = [
        ["NS-VICTOR", "2025-01-15 08:00", "sunflower", 25000, 15000, "HIGH", "Yes", "Example sunflower vessel"],
        ["MV-PALMOIL1", "2025-01-18 14:00", "palm", 18000, 12000, "MEDIUM", "Yes", "Example palm vessel"],
    ]
    
    for row_idx, vessel in enumerate(sample_vessels, 2):
        for col_idx, value in enumerate(vessel, 1):
            cell = ws_vessels.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            if col_idx == 2:  # ETA column
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    # Add empty rows for user input
    for row_idx in range(4, 11):
        for col_idx in range(1, len(headers) + 1):
            cell = ws_vessels.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # Column widths
    ws_vessels.column_dimensions['A'].width = 15
    ws_vessels.column_dimensions['B'].width = 18
    ws_vessels.column_dimensions['C'].width = 12
    ws_vessels.column_dimensions['D'].width = 14
    ws_vessels.column_dimensions['E'].width = 15
    ws_vessels.column_dimensions['F'].width = 10
    ws_vessels.column_dimensions['G'].width = 8
    ws_vessels.column_dimensions['H'].width = 30
    
    # ========================================
    # Sheet 2: Input_CargoPlan (for palm vessels)
    # ========================================
    ws_cargo = wb.create_sheet("Input_CargoPlan")
    
    cargo_headers = ["VesselID", "Hold", "Fraction", "Volume", "Notes"]
    for col_idx, header in enumerate(cargo_headers, 1):
        cell = ws_cargo.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Sample cargo plan
    sample_cargo = [
        ["MV-PALMOIL1", "H1", "Stearin", 5000, "Hard fraction"],
        ["MV-PALMOIL1", "H2", "Olein", 8000, "Liquid fraction"],
        ["MV-PALMOIL1", "H3", "LOW3MCPD", 5000, "Premium grade"],
    ]
    
    for row_idx, cargo in enumerate(sample_cargo, 2):
        for col_idx, value in enumerate(cargo, 1):
            cell = ws_cargo.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
    
    # Empty rows
    for row_idx in range(5, 21):
        for col_idx in range(1, len(cargo_headers) + 1):
            cell = ws_cargo.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    ws_cargo.column_dimensions['A'].width = 15
    ws_cargo.column_dimensions['B'].width = 8
    ws_cargo.column_dimensions['C'].width = 15
    ws_cargo.column_dimensions['D'].width = 12
    ws_cargo.column_dimensions['E'].width = 30
    
    # ========================================
    # Sheet 3: Input_CurrentState
    # ========================================
    ws_state = wb.create_sheet("Input_CurrentState")
    
    state_headers = ["TankID", "CurrentVolume", "CurrentProduct", "State", "Notes"]
    for col_idx, header in enumerate(state_headers, 1):
        cell = ws_state.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Default tank state (empty)
    tank_state = [
        ["RVS-1", 0, "-", "idle", "9900t capacity, sunflower only"],
        ["RVS-2", 0, "-", "idle", "9900t capacity, sunflower only"],
        ["RVS-3", 0, "-", "idle", "9900t capacity, palm only"],
        ["RVS-5", 0, "-", "idle", "2700t capacity, swing tank"],
        ["SHT", 0, "-", "idle", "2700t capacity, shore tank"],
    ]
    
    for row_idx, tank in enumerate(tank_state, 2):
        for col_idx, value in enumerate(tank, 1):
            cell = ws_state.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
    
    ws_state.column_dimensions['A'].width = 10
    ws_state.column_dimensions['B'].width = 15
    ws_state.column_dimensions['C'].width = 18
    ws_state.column_dimensions['D'].width = 12
    ws_state.column_dimensions['E'].width = 35
    
    # ========================================
    # Sheet 4: Input_Rail
    # ========================================
    ws_rail = wb.create_sheet("Input_Rail")
    
    rail_headers = ["Date", "Product", "Volume", "Batches", "Notes"]
    for col_idx, header in enumerate(rail_headers, 1):
        cell = ws_rail.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Sample rail schedule
    base_date = datetime(2025, 1, 10)
    sample_rail = [
        [(base_date + timedelta(days=i)).strftime("%Y-%m-%d"), "sunflower", 1170 * (i%3 + 1), (i%3 + 1), f"Day {i+1}"]
        for i in range(5)
    ]
    
    for row_idx, rail in enumerate(sample_rail, 2):
        for col_idx, value in enumerate(rail, 1):
            cell = ws_rail.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            if col_idx == 1:  # Date column
                cell.number_format = 'YYYY-MM-DD'
    
    # Empty rows
    for row_idx in range(7, 21):
        for col_idx in range(1, len(rail_headers) + 1):
            cell = ws_rail.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    ws_rail.column_dimensions['A'].width = 12
    ws_rail.column_dimensions['B'].width = 15
    ws_rail.column_dimensions['C'].width = 12
    ws_rail.column_dimensions['D'].width = 10
    ws_rail.column_dimensions['E'].width = 30
    
    # ========================================
    # Sheet 5: Input_Demand (Optional)
    # ========================================
    ws_demand = wb.create_sheet("Input_Demand")
    
    demand_headers = ["Week", "Client", "Product", "Volume", "Priority", "Notes"]
    for col_idx, header in enumerate(demand_headers, 1):
        cell = ws_demand.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Sample demand
    sample_demand = [
        ["2025-W03", "Client A", "Olein", 5000, "HIGH", "Strategic client"],
        ["2025-W03", "Client B", "Stearin", 3000, "MEDIUM", "Regular client"],
    ]
    
    for row_idx, demand in enumerate(sample_demand, 2):
        for col_idx, value in enumerate(demand, 1):
            cell = ws_demand.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
    
    # Empty rows
    for row_idx in range(4, 21):
        for col_idx in range(1, len(demand_headers) + 1):
            cell = ws_demand.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    ws_demand.column_dimensions['A'].width = 12
    ws_demand.column_dimensions['B'].width = 15
    ws_demand.column_dimensions['C'].width = 15
    ws_demand.column_dimensions['D'].width = 12
    ws_demand.column_dimensions['E'].width = 10
    ws_demand.column_dimensions['F'].width = 30
    
    # ========================================
    # Sheet 6: Settings
    # ========================================
    ws_settings = wb.create_sheet("Settings")
    
    settings_headers = ["Parameter", "Value", "Description"]
    for col_idx, header in enumerate(settings_headers, 1):
        cell = ws_settings.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    settings_data = [
        ["Run Mode", "All", "Options: All, Palm Only, Sunflower Only"],
        ["Optimize Sequence", "No", "Options: Yes, No"],
        ["Vessel Filter", "", "Enter Vessel ID to run single simulation"],
    ]
    
    for row_idx, setting in enumerate(settings_data, 2):
        for col_idx, value in enumerate(setting, 1):
            cell = ws_settings.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 2: # Value column
                cell.fill = example_fill
                
    ws_settings.column_dimensions['A'].width = 20
    ws_settings.column_dimensions['B'].width = 20
    ws_settings.column_dimensions['C'].width = 40

    # ========================================
    # Sheet 7: Instructions
    # ========================================
    ws_instructions = wb.create_sheet("Instructions", 0)  # First sheet
    
    instructions_text = [
        ["TERMINAL OPTIMIZER - QUICK START", "", "", ""],
        ["", "", "", ""],
        ["Step 1: Fill Input Sheets", "", "", ""],
        ["  • Input_Vessels: Add your vessel schedule", "", "", ""],
        ["  • Input_CargoPlan: For palm vessels, specify holds and fractions", "", "", ""],
        ["  • Input_CurrentState: Set current tank inventory (default is empty)", "", "", ""],
        ["  • Input_Rail: Add railway deliveries (optional)", "", "", ""],
        ["  • Input_Demand: Add client orders (optional)", "", "", ""],
        ["", "", "", ""],
        ["Step 2: Save Your File", "", "", ""],
        ["  • Click Save As", "", "", ""],
        ["  • Save as: my_scenario.xlsx", "", "", ""],
        ["", "", "", ""],
        ["Step 3: Run Calculation", "", "", ""],
        ["  • Open Command Prompt in installation folder", "", "", ""],
        ["  • Type: run.bat my_scenario.xlsx", "", "", ""],
        ["  • Press Enter", "", "", ""],
        ["", "", "", ""],
        ["Step 4: Review Results", "", "", ""],
        ["  • Program creates: my_scenario_optimized.xlsx", "", "", ""],
        ["  • Check Output_Operations for operation timeline", "", "", ""],
        ["  • Check Output_Balances for tank levels", "", "", ""],
        ["  • Check Output_Summary for key metrics", "", "", ""],
        ["", "", "", ""],
        ["Blue rows are EXAMPLES - you can modify or delete them", "", "", ""],
        ["White rows are empty - add your own data here", "", "", ""],
        ["", "", "", ""],
        ["For detailed help, see:", "", "", ""],
        ["  • QUICKSTART_EXCEL.md - Step-by-step guide", "", "", ""],
        ["  • docs/USER_MANUAL.md - Complete reference", "", "", ""],
        ["  • INSTALLATION_GUIDE.md - Installation help", "", "", ""],
    ]
    
    for row_idx, row in enumerate(instructions_text, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.fill = header_fill
                cell.font = Font(bold=True, size=14, color="FFFFFF")
            elif "Step" in str(value):
                cell.font = Font(bold=True, size=12)
            elif value.startswith("  •"):
                cell.font = Font(size=10)
    
    ws_instructions.column_dimensions['A'].width = 60
    ws_instructions.merge_cells('A1:D1')
    
    # Save workbook
    wb.save(filename)
    print(f"[OK] Created Excel template: {filename}")
    print( "[OK] Sample data included (highlighted in blue)")
    print("[OK] Ready to use with Terminal Optimizer!")
    
    return filename

if __name__ == "__main__":
    create_template_with_sample_data()
