"""
Visualization Module for Terminal Optimizer

This module handles the generation of visual representations of the schedule,
primarily focusing on Excel-based Gantt charts.
"""

from __future__ import annotations
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from terminal_optimizer.domain_models import Operation, OperationType

class ExcelGanttChart:
    """Generates a Gantt chart in Excel format."""
    
    def __init__(self, writer: pd.ExcelWriter):
        """
        Initialize with an existing ExcelWriter.
        
        Args:
            writer: pandas ExcelWriter object (engine='openpyxl')
        """
        self.writer = writer
        self.workbook = writer.book
        
    def create_chart(self, operations: List[Operation], start_date: Optional[datetime] = None, days: int = 30):
        """
        Create a Gantt chart sheet in the workbook.
        
        Args:
            operations: List of operations to visualize
            start_date: Start date for the timeline (default: earliest operation start)
            days: Number of days to visualize
        """
        if not operations:
            return
            
        # Determine start date if not provided
        if start_date is None:
            valid_starts = [op.start_time for op in operations if op.start_time]
            if valid_starts:
                start_date = min(valid_starts).replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Create sheet
        ws = self.workbook.create_sheet("Gantt_Chart")
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Operation colors
        colors = {
            OperationType.BERTHING: "FFC000",   # Orange
            OperationType.UNBERTHING: "FFC000", # Orange
            OperationType.DISCHARGE: "92D050",  # Green
            OperationType.LOAD: "00B050",       # Dark Green
            OperationType.TRANSFER: "00B0F0",   # Blue
            OperationType.CLEANING: "FF0000",   # Red
            OperationType.ANALYSIS: "7030A0",   # Purple
            OperationType.RAIL_BATCH: "C00000", # Dark Red
        }
        
        # Resources to track
        resources = [
            "BERTH",
            "Line1",
            "Line2",
            "RVS-1",
            "RVS-2",
            "RVS-3",
            "RVS-5",
            "SHT",
            "Rail"
        ]
        
        # Setup headers
        ws.cell(row=1, column=1, value="Resource").fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        
        # Time headers (every 6 hours)
        col_idx = 2
        current_time = start_date
        end_time = start_date + timedelta(days=days)
        
        time_map = {}  # Map datetime to column index
        
        while current_time <= end_time:
            cell = ws.cell(row=1, column=col_idx, value=current_time.strftime("%d-%H"))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", text_rotation=90)
            ws.column_dimensions[get_column_letter(col_idx)].width = 4
            
            time_map[current_time] = col_idx
            current_time += timedelta(hours=6)
            col_idx += 1
            
        # Setup resource rows
        resource_rows = {}
        for i, resource in enumerate(resources, start=2):
            cell = ws.cell(row=i, column=1, value=resource)
            cell.font = Font(bold=True)
            resource_rows[resource] = i
            
        # Plot operations
        for op in operations:
            if not op.start_time or not op.end_time:
                continue
                
            # Determine resources used
            used_resources = []
            
            # Berth usage
            if op.op_type in [OperationType.BERTHING, OperationType.UNBERTHING, OperationType.LOAD, OperationType.DISCHARGE]:
                if op.vessel_id:
                    used_resources.append("BERTH")
            
            # Line usage
            if op.line:
                used_resources.append(op.line)
                
            # Tank usage
            if op.source in resource_rows:
                used_resources.append(op.source)
            if op.target in resource_rows:
                used_resources.append(op.target)
                
            # Rail usage
            if "Rail" in op.source or "Rail" in op.target:
                used_resources.append("Rail")
                
            # Color code
            color = colors.get(op.op_type, "CCCCCC")
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Plot on timeline
            start_col = self._get_time_column(op.start_time, start_date, time_map)
            end_col = self._get_time_column(op.end_time, start_date, time_map)
            
            for res in used_resources:
                if res not in resource_rows:
                    continue
                    
                row = resource_rows[res]
                
                # Fill cells
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    
                    # Add comment or value for details (optional, keeping it simple for now)
                    if col == start_col:
                        cell.value = op.op_id
    
    def _get_time_column(self, time: datetime, start_date: datetime, time_map: Dict[datetime, int]) -> int:
        """Map a time to the nearest column index."""
        # Round to nearest 6 hours
        delta = time - start_date
        hours = delta.total_seconds() / 3600
        six_hour_blocks = int(hours / 6)
        
        # Calculate expected column (2 is start column)
        col = 2 + six_hour_blocks
        return max(2, col)
